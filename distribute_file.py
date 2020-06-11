from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook
from shutil import copy
import os

def select_student_list():
    root.class_list = filedialog.askopenfilename(initialdir = ".", title = "Select Student List", filetypes=(("Excel files","*.xlsx"),("All files","*.*")))
    student_list_lb.config(text=os.path.basename(root.class_list))

def select_source_file():
    root.source_file = filedialog.askopenfilename(initialdir = ".", title = "Select Source File")
    source_file_lb.config(text=os.path.basename(root.source_file))

def select_destination():
    root.destination = filedialog.askdirectory(initialdir = ".", title = "Select Desintation Folder")
    destination_folder_lb.config(text=root.destination)


def go():
    # create student list
    workbook = load_workbook(filename=root.class_list)
    sheet = workbook.active
    row_count = 3
    while sheet.cell(row=row_count, column=1).value != None:
        # copy file to new location
        copy(root.source_file,root.destination)
        # create new file name
        file_name = os.path.basename(root.source_file)
        prefix = sheet.cell(row=row_count, column=1).value +\
            " - " + sheet.cell(row=row_count, column=2).value +\
            ", " + sheet.cell(row=row_count, column=3).value + " - "
        old_file_name = root.destination+"/"+file_name
        new_file_name = root.destination+"/"+prefix+file_name
        # rename file
        os.rename(old_file_name,new_file_name)

        row_count += 1
    go_lb.config(text="Done")

root = Tk()
root.geometry("600x400")
root.title("CaRAF")

# create file variables
root.class_list = ""
root.source_file = ""
root.destination = ""



Label(root, text="Copy and Rename Assessment Files", font=("Arial",20))\
    .grid(row=0,column=0,columnspan=2)

Button(root, text="Select student list", width = 20, command=select_student_list)\
    .grid(row=1,column=0,padx=5, pady=5)
student_list_lb = Label(root)
student_list_lb.grid(row=1,column=1, sticky="W")

Button(root, text="Select source file", width = 20, command=select_source_file)\
    .grid(row=2,column=0,padx=5, pady=5)
source_file_lb = Label(root)
source_file_lb.grid(row=2,column=1, sticky="W")

Button(root, text="Select destination folder", width = 20, command=select_destination)\
    .grid(row=3,column=0,padx=5, pady=5)
destination_folder_lb = Label(root)
destination_folder_lb.grid(row=3,column=1, sticky="W")

Button(root,text="Go", width = 20,command=go)\
    .grid(row=4,column=0,padx=5, pady=5)
go_lb = Label(root)
go_lb.grid(row=4, column=1, sticky="W")


root.mainloop()